import openpyxl
from pymongo import MongoClient

client = MongoClient('mongodb://localhost:27017/')
db = client['attendance_system']

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        class_name = sheet
        sheet_data = workbook[sheet]
        students = []
        for row in sheet_data.iter_rows(min_row=2, values_only=True):
            if len(row) == 5:
                roll_number, name, semester, year, section= row
            elif len(row) == 4:
                roll_number, name, semester, year = row
                section = 'Nil'  
            else:
                continue  

            if not db.students.find_one({'roll_number': roll_number, 'name': name, 'class': class_name, 'section': section}):
                students.append({
                    'roll_number': roll_number,
                    'name': name,
                    'class': class_name,
                    'year': year,
                    'semester': semester,
                    'section': section
                })
        if students:
            db.students.insert_many(students)

read_excel('2nd_std_lst.xlsx')
