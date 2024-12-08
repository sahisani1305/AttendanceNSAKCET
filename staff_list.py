import openpyxl
from pymongo import MongoClient

client = MongoClient('mongodb://localhost:27017/')
db = client['attendance_system']

def read_excel_for_staff(file_path):
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        sheet_data = workbook[sheet]
        staff_members = []
        for row in sheet_data.iter_rows(min_row=2, values_only=True):
            if len(row) == 2:
                staff_id, name, = row
            else:
                continue  

            if not db.staff.find_one({'staff_id': staff_id, 'name': name}):
                staff_members.append({
                    'staff_id': staff_id,
                    'name': name,
                })
        if staff_members:
            db.staff.insert_many(staff_members)

read_excel_for_staff('staff_list.xlsx')
